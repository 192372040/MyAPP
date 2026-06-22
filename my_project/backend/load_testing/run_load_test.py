"""
MediConnect – Load Test Runner
================================
Runs Locust in headless mode (100 users, 1 min) against localhost:5000,
then converts the CSV output into a formatted Excel report.

Usage:
    python run_load_test.py
    python run_load_test.py --host http://192.168.1.10:5000
    python run_load_test.py --users 200 --duration 2m

Output files (inside results/ folder):
    load_test_stats.csv          – Per-endpoint request statistics
    load_test_stats_history.csv  – Time-series throughput/response data
    load_test_failures.csv       – Any failed requests
    MediConnect_Load_Test_Report.xlsx  – Formatted Excel workbook
"""

import subprocess
import sys
import os
import csv
import argparse
from datetime import datetime
from pathlib import Path

# ── Optional: openpyxl for Excel export ──────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed – Excel export will be skipped.")
    print("       Run:  pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
RESULTS_DIR  = SCRIPT_DIR / "results"
LOCUST_FILE  = SCRIPT_DIR / "locustfile.py"
CSV_PREFIX   = str(RESULTS_DIR / "load_test")
EXCEL_OUTPUT = RESULTS_DIR / "MediConnect_Load_Test_Report.xlsx"

BRAND_COLOR  = "1E3A5F"   # dark navy
ACCENT_COLOR = "2E86AB"   # blue
PASS_COLOR   = "27AE60"   # green
WARN_COLOR   = "F39C12"   # amber
FAIL_COLOR   = "E74C3C"   # red
HEADER_FONT  = "Calibri"
BODY_FONT    = "Calibri"

# ─────────────────────────────────────────────────────────────────────────────


def parse_args():
    p = argparse.ArgumentParser(description="MediConnect Load Test Runner")
    p.add_argument("--host",     default="http://localhost:5000",
                   help="Backend host URL")
    p.add_argument("--users",    default="100",
                   help="Number of virtual users (default: 100)")
    p.add_argument("--spawn",    default="10",
                   help="Spawn rate – users per second (default: 10)")
    p.add_argument("--duration", default="1m",
                   help="Test duration, e.g. 1m, 90s (default: 1m)")
    return p.parse_args()


# ─────────────────────────────────────────────────────────────────────────────
# 1.  RUN LOCUST
# ─────────────────────────────────────────────────────────────────────────────

def run_locust(host: str, users: str, spawn: str, duration: str):
    RESULTS_DIR.mkdir(exist_ok=True)

    cmd = [
        sys.executable, "-m", "locust",
        "-f", str(LOCUST_FILE),
        "--headless",
        "--host",     host,
        "-u",         users,
        "-r",         spawn,
        "--run-time", duration,
        "--csv",      CSV_PREFIX,
        "--csv-full-history",
        "--only-summary",
    ]

    print("\n" + "-" * 60)
    print("  MediConnect Baseline / Load Test")
    print("-" * 60)
    print(f"  Host     : {host}")
    print(f"  Users    : {users}")
    print(f"  Spawn    : {spawn} users/sec")
    print(f"  Duration : {duration}")
    print(f"  CSV out  : {RESULTS_DIR}")
    print("-" * 60 + "\n")

    result = subprocess.run(cmd, capture_output=False, text=True)
    if result.returncode not in (0, 1):
        print(f"\n[ERROR] Locust exited with unexpected code {result.returncode}")
        sys.exit(result.returncode)
    elif result.returncode == 1:
        print("\n[WARN] Locust exited with code 1 (failures detected). Continuing to generate report...\n")

    print("\n[OK] Locust finished.\n")


# ─────────────────────────────────────────────────────────────────────────────
# 2.  READ CSVs
# ─────────────────────────────────────────────────────────────────────────────

def read_csv(path: str) -> list[dict]:
    p = Path(path)
    if not p.exists():
        print(f"[WARN] CSV not found: {p}")
        return []
    with open(p, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ─────────────────────────────────────────────────────────────────────────────
# 3.  HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, name=BODY_FONT) -> Font:
    return Font(bold=bold, color=color, size=size, name=name)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fmt_ms(val) -> str:
    try:
        return f"{float(val):.1f} ms"
    except (TypeError, ValueError):
        return str(val)


def _fmt_rps(val) -> str:
    try:
        return f"{float(val):.2f} req/s"
    except (TypeError, ValueError):
        return str(val)


def _grade(avg_ms: float, fail_pct: float) -> tuple[str, str]:
    """Return (grade, hex_color) based on avg response time and failure %."""
    if fail_pct > 5 or avg_ms > 20000:
        return "POOR", FAIL_COLOR
    if fail_pct > 1 or avg_ms > 15000:
        return "DEGRADED", WARN_COLOR
    return "PASS", PASS_COLOR


# ─────────────────────────────────────────────────────────────────────────────
# 4.  BUILD EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(stats: list[dict], history: list[dict], failures: list[dict],
                meta: dict):
    wb = openpyxl.Workbook()

    _sheet_summary(wb, stats, meta)
    _sheet_per_endpoint(wb, stats)
    _sheet_history(wb, history)
    _sheet_failures(wb, failures)

    wb.save(EXCEL_OUTPUT)
    print(f"[OK] Excel report saved -> {EXCEL_OUTPUT}\n")


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _sheet_summary(wb, stats: list[dict], meta: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "🏥  MediConnect – Baseline Load Test Report"
    c.font  = Font(bold=True, size=16, color="FFFFFF", name=HEADER_FONT)
    c.fill  = _fill(BRAND_COLOR)
    c.alignment = _center()
    ws.row_dimensions[1].height = 34

    # Sub-title row
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Generated: {meta['timestamp']}   |   Host: {meta['host']}   |   Users: {meta['users']}   |   Duration: {meta['duration']}"
    c.font  = Font(italic=True, size=10, color="FFFFFF", name=BODY_FONT)
    c.fill  = _fill(ACCENT_COLOR)
    c.alignment = _center()
    ws.row_dimensions[2].height = 20

    # Blank row
    ws.row_dimensions[3].height = 8

    # KPI boxes  (row 4–6)
    aggregated = _get_aggregated(stats)
    kpis = [
        ("Total Requests",      aggregated.get("total_req",   "—")),
        ("Requests / Second",   aggregated.get("rps",         "—")),
        ("Avg Response Time",   aggregated.get("avg_ms",      "—")),
        ("Min Response Time",   aggregated.get("min_ms",      "—")),
        ("Max Response Time",   aggregated.get("max_ms",      "—")),
        ("Failure Rate",        aggregated.get("fail_pct",    "—")),
        ("Overall Grade",       aggregated.get("grade",       "—")),
    ]

    col_map = ["A", "B", "C", "D", "E", "F", "G"]
    col_widths = [18, 18, 20, 20, 20, 16, 16]
    grade_color = aggregated.get("grade_color", PASS_COLOR)

    for i, ((label, value), col) in enumerate(zip(kpis, col_map)):
        ws.merge_cells(f"{col}4:{col}5")
        ws[f"{col}4"].value    = label
        ws[f"{col}4"].font     = Font(bold=True, size=9, color="FFFFFF", name=BODY_FONT)
        ws[f"{col}4"].fill     = _fill(ACCENT_COLOR)
        ws[f"{col}4"].alignment = _center()

        ws.merge_cells(f"{col}6:{col}7")
        cell = ws[f"{col}6"]
        cell.value     = value
        cell.font      = Font(bold=True, size=13, name=HEADER_FONT,
                              color="FFFFFF" if i == 6 else "1E3A5F")
        cell.alignment = _center()
        if i == 6:
            cell.fill = _fill(grade_color)
        else:
            cell.fill = _fill("EBF5FB")
        cell.border = _border()
        ws.column_dimensions[col].width = col_widths[i]

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 28

    # Blank row
    ws.row_dimensions[8].height = 10

    # Section header
    ws.merge_cells("A9:G9")
    c = ws["A9"]
    c.value     = "Performance Thresholds"
    c.font      = Font(bold=True, size=11, color="FFFFFF", name=HEADER_FONT)
    c.fill      = _fill(BRAND_COLOR)
    c.alignment = _center()
    ws.row_dimensions[9].height = 20

    thresholds = [
        ("Metric",              "Target",           "Actual",                   "Status"),
        ("Avg Response Time",   "< 15000 ms",       aggregated.get("avg_ms","—"),
         "\u2713" if _num(aggregated.get("avg_ms","99999")) < 15000 else "\u2717"),
        ("Max Response Time",   "< 20000 ms",       aggregated.get("max_ms","—"),
         "\u2713" if _num(aggregated.get("max_ms","99999")) < 20000 else "\u2717"),
        ("Failure Rate",        "< 1 %",            aggregated.get("fail_pct","—"),
         "\u2713" if _num(aggregated.get("fail_pct","100")) < 1 else "\u2717"),
        ("Requests / Second",   "> 5 req/s",        aggregated.get("rps","—"),
         "\u2713" if _num(aggregated.get("rps","0")) > 5 else "\u2717"),
    ]

    for r_idx, row in enumerate(thresholds, start=10):
        ws.row_dimensions[r_idx].height = 18
        colors = [BRAND_COLOR, BRAND_COLOR, ACCENT_COLOR, ACCENT_COLOR]
        is_header = r_idx == 10
        for c_idx, (col, val) in enumerate(zip(col_map[:4], row)):
            cell = ws.cell(row=r_idx, column=c_idx + 1, value=val)
            cell.alignment = _center()
            cell.border    = _border()
            if is_header:
                cell.font = Font(bold=True, color="FFFFFF", size=10, name=HEADER_FONT)
                cell.fill = _fill(ACCENT_COLOR)
            else:
                cell.font = Font(size=10, name=BODY_FONT)
                cell.fill = _fill("F2F9FF")
                if c_idx == 3:  # Status column
                    cell.fill = _fill(PASS_COLOR if val == "✓" else FAIL_COLOR)
                    cell.font = Font(bold=True, color="FFFFFF", size=12, name=BODY_FONT)


def _get_aggregated(stats: list[dict]) -> dict:
    """Compute aggregate KPIs from stats rows (excluding the Aggregated row)."""
    rows = [r for r in stats if r.get("Name", "") != "Aggregated"]
    agg  = next((r for r in stats if r.get("Name", "") == "Aggregated"), {})

    total_req  = int(agg.get("Request Count", 0) or 0)
    total_fail = int(agg.get("Failure Count", 0) or 0)
    fail_pct   = (total_fail / total_req * 100) if total_req else 0
    avg_ms     = float(agg.get("Average Response Time", 0) or 0)
    min_ms     = float(agg.get("Min Response Time", 0) or 0)
    max_ms     = float(agg.get("Max Response Time", 0) or 0)
    rps        = float(agg.get("Requests/s", 0) or 0)
    grade, gc  = _grade(avg_ms, fail_pct)

    return {
        "total_req":  f"{total_req:,}",
        "rps":        _fmt_rps(rps),
        "avg_ms":     _fmt_ms(avg_ms),
        "min_ms":     _fmt_ms(min_ms),
        "max_ms":     _fmt_ms(max_ms),
        "fail_pct":   f"{fail_pct:.2f} %",
        "grade":      grade,
        "grade_color": gc,
    }


def _num(s: str) -> float:
    try:
        return float("".join(c for c in str(s) if c in "0123456789."))
    except ValueError:
        return 0.0


# ── Sheet 2: Per-Endpoint ─────────────────────────────────────────────────────

def _sheet_per_endpoint(wb, stats: list[dict]):
    ws = wb.create_sheet("Per-Endpoint Stats")
    ws.sheet_view.showGridLines = False

    headers = [
        "Endpoint", "Method", "Requests", "Failures",
        "Fail %", "Avg (ms)", "Min (ms)", "Max (ms)",
        "Median (ms)", "90th % (ms)", "95th % (ms)", "99th % (ms)",
        "Req/s", "Grade",
    ]

    col_widths = [35, 8, 11, 10, 8, 10, 10, 10, 11, 12, 12, 12, 10, 12]

    # Header
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name=HEADER_FONT)
        cell.fill      = _fill(BRAND_COLOR)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[1].height = 20

    for r_idx, row in enumerate(stats, start=2):
        req   = int(row.get("Request Count", 0) or 0)
        fail  = int(row.get("Failure Count", 0) or 0)
        fp    = (fail / req * 100) if req else 0
        avg   = float(row.get("Average Response Time", 0) or 0)
        grade, gc = _grade(avg, fp)

        row_fill = _fill("FAFEFE") if r_idx % 2 == 0 else _fill("FFFFFF")

        values = [
            row.get("Name", ""),
            row.get("Type", ""),
            req,
            fail,
            f"{fp:.1f}%",
            round(avg, 1),
            round(float(row.get("Min Response Time", 0) or 0), 1),
            round(float(row.get("Max Response Time", 0) or 0), 1),
            round(float(row.get("Median Response Time", 0) or 0), 1),
            round(float(row.get("90%", 0) or 0), 1),
            round(float(row.get("95%", 0) or 0), 1),
            round(float(row.get("99%", 0) or 0), 1),
            round(float(row.get("Requests/s", 0) or 0), 2),
            grade,
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = _center() if c_idx > 1 else _left()
            cell.border    = _border()
            cell.font      = Font(size=10, name=BODY_FONT)
            if c_idx == len(values):   # Grade column
                cell.fill = _fill(gc)
                cell.font = Font(bold=True, color="FFFFFF", size=10, name=BODY_FONT)
            else:
                cell.fill = row_fill
        ws.row_dimensions[r_idx].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 3: History (time series) ───────────────────────────────────────────

def _sheet_history(wb, history: list[dict]):
    ws = wb.create_sheet("Throughput History")
    ws.sheet_view.showGridLines = False

    if not history:
        ws["A1"] = "No history data available."
        return

    headers = ["Timestamp", "User Count", "Req/s", "Failures/s",
               "Avg Response (ms)", "Min Response (ms)", "Max Response (ms)"]
    col_widths = [22, 12, 12, 14, 20, 20, 20]

    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name=HEADER_FONT)
        cell.fill      = _fill(ACCENT_COLOR)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[1].height = 20

    # Only keep "Aggregated" rows for the history view
    agg_rows = [r for r in history if r.get("Name", "") == "Aggregated"]

    for r_idx, row in enumerate(agg_rows, start=2):
        values = [
            row.get("Timestamp", ""),
            row.get("User Count", ""),
            round(float(row.get("Requests/s", 0) or 0), 2),
            round(float(row.get("Failures/s", 0) or 0), 2),
            round(float(row.get("Total Average Response Time", 0) or 0), 1),
            round(float(row.get("Total Min Response Time", 0) or 0), 1),
            round(float(row.get("Total Max Response Time", 0) or 0), 1),
        ]
        row_fill = _fill("F2F9FF") if r_idx % 2 == 0 else _fill("FFFFFF")
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = _center()
            cell.border    = _border()
            cell.fill      = row_fill
            cell.font      = Font(size=10, name=BODY_FONT)
        ws.row_dimensions[r_idx].height = 15

    ws.freeze_panes = "A2"

    # ── Embedded bar chart: Req/s over time ──
    if len(agg_rows) > 1:
        last_data_row = len(agg_rows) + 1
        chart = BarChart()
        chart.type  = "col"
        chart.title = "Requests/s over Time"
        chart.y_axis.title = "Req/s"
        chart.x_axis.title = "Time point"
        chart.style = 10
        chart.width  = 22
        chart.height = 12

        data = Reference(ws, min_col=3, min_row=1,
                         max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "I3")


# ── Sheet 4: Failures ─────────────────────────────────────────────────────────

def _sheet_failures(wb, failures: list[dict]):
    ws = wb.create_sheet("Failures")
    ws.sheet_view.showGridLines = False

    if not failures:
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value     = "✅  No failures recorded during this test run."
        c.font      = Font(bold=True, size=12, color=PASS_COLOR, name=BODY_FONT)
        c.alignment = _center()
        c.fill      = _fill("EAFAF1")
        ws.column_dimensions["A"].width = 60
        return

    headers    = ["Method", "Endpoint", "Error", "Occurrences"]
    col_widths = [10, 40, 60, 14]

    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name=HEADER_FONT)
        cell.fill      = _fill(FAIL_COLOR)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[1].height = 20

    for r_idx, row in enumerate(failures, start=2):
        values = [
            row.get("Method", ""),
            row.get("Name", ""),
            row.get("Error", ""),
            int(row.get("Occurrences", 0) or 0),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = _left() if c_idx == 3 else _center()
            cell.border    = _border()
            cell.fill      = _fill("FFF0F0")
            cell.font      = Font(size=10, name=BODY_FONT)
        ws.row_dimensions[r_idx].height = 16

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# 5.  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()

    # Step 1 – Run test
    run_locust(args.host, args.users, args.spawn, args.duration)

    # Step 2 – Read CSVs
    stats    = read_csv(f"{CSV_PREFIX}_stats.csv")
    history  = read_csv(f"{CSV_PREFIX}_stats_history.csv")
    failures = read_csv(f"{CSV_PREFIX}_failures.csv")

    # Step 3 – Build Excel
    if HAS_OPENPYXL:
        meta = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "host":      args.host,
            "users":     args.users,
            "duration":  args.duration,
        }
        build_excel(stats, history, failures, meta)
    else:
        print("[INFO] Install openpyxl to generate the Excel report:")
        print("       pip install openpyxl")

    # Step 4 – Console summary
    _print_console_summary(stats)


def _print_console_summary(stats: list[dict]):
    agg = next((r for r in stats if r.get("Name") == "Aggregated"), {})
    if not agg:
        return

    req    = int(agg.get("Request Count", 0) or 0)
    fail   = int(agg.get("Failure Count", 0) or 0)
    avg_ms = float(agg.get("Average Response Time", 0) or 0)
    min_ms = float(agg.get("Min Response Time", 0) or 0)
    max_ms = float(agg.get("Max Response Time", 0) or 0)
    rps    = float(agg.get("Requests/s", 0) or 0)
    fp     = (fail / req * 100) if req else 0

    print("-" * 60)
    print("  LOAD TEST RESULTS")
    print("-" * 60)
    print(f"  Total Requests   : {req:,}")
    print(f"  Requests/sec     : {rps:.2f} req/s")
    print(f"  Failures         : {fail}  ({fp:.2f}%)")
    print(f"  Avg Response     : {avg_ms:.1f} ms")
    print(f"  Min Response     : {min_ms:.1f} ms")
    print(f"  Max Response     : {max_ms:.1f} ms")
    print("-" * 60)
    print(f"  Excel Report     : {EXCEL_OUTPUT}")
    print("-" * 60 + "\n")


if __name__ == "__main__":
    main()
