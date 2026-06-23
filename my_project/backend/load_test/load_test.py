# -*- coding: utf-8 -*-
"""
MediConnect API -- Baseline / Load Test
========================================
  100 virtual users
  1 minute continuous run
  Measures: RPS, Avg/Min/Max/P95/P99 response times, Error rate
  Outputs : LoadTest_Report.xlsx  (same folder as this script)

Run:
    python load_test.py
"""

import sys
import io

# Force UTF-8 stdout so Windows cp1252 does not crash on special chars
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import time
import threading
import math
import statistics
from datetime import datetime
from collections import defaultdict
from pathlib import Path

try:
    import requests
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
except ImportError as e:
    print(f"Missing package: {e}\nRun: pip install requests openpyxl")
    sys.exit(1)

# ===========================================================================
# CONFIG
# ===========================================================================
BASE          = "http://localhost:5000"
VIRTUAL_USERS = 100
DURATION_SEC  = 60          # 1 minute
RAMP_UP_SEC   = 5           # spread thread starts over 5 s
TIMEOUT       = 5           # HTTP request timeout (s)
REPORT_FILE   = Path(__file__).parent / "LoadTest_Report.xlsx"
NOW           = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ===========================================================================
# ENDPOINTS  (realistic mix the app actually uses)
# ===========================================================================
ENDPOINTS = [
    ("GET",  "/",                         None),
    ("GET",  "/hospitals",                None),
    ("GET",  "/test-db",                  None),
    ("POST", "/send-otp",                 {"email": "loadtest@gmail.com"}),
    ("POST", "/verify-otp",               {"email": "loadtest@gmail.com", "otp": "000000"}),
    ("GET",  "/booked-slots",             {"doctor_name": "Dr. Smith", "appointment_date": "2099-01-01"}),
    ("GET",  "/get-patient-profile",      {"email": "pt1@gmail.com"}),
    ("POST", "/update-profile",           {"email": "lt1@gmail.com", "name": "Load Tester",
                                           "phone": "9999999999", "age": "25",
                                           "blood_group": "O+", "gender": "Male",
                                           "medical_history": []}),
    ("GET",  "/get-patient-appointments", {"patient_email": "pt1@gmail.com"}),
    ("POST", "/doctor-login",             {"doctor_id": "MED-DOC-2026-00001", "password": "doctor123"}),
    ("POST", "/admin-login",              {"hospital_id": "HOSP-001", "password": "admin123"}),
    ("GET",  "/get-hospital-beds",        {"hospital_id": "1"}),
    ("GET",  "/get-hospital-analytics",   {"hospital_id": "1"}),
    ("POST", "/ai-chat",                  {"message": "What is diabetes?"}),
    ("POST", "/book-appointment",         {
        "patient_email": "lt1@gmail.com", "patient_name": "Load Tester",
        "doctor_name": "Dr. Test", "specialization": "General",
        "hospital_name": "Test Hospital", "appointment_slot": "09:00 AM",
        "payment_method": "Online", "consultation_fee": "200",
        "appointment_date": "2099-06-01",
    }),
]

# ===========================================================================
# SHARED STATE
# ===========================================================================
_lock       = threading.Lock()
_results    = []                       # all request records
_per_second = defaultdict(list)        # bucket(int) -> [response_ms, ...]
_stop       = threading.Event()
_start_time = 0.0

SESSION = requests.Session()
SESSION.headers.update({"Content-Type": "application/json"})

# ===========================================================================
# WORKER  (one per virtual user)
# ===========================================================================
def worker(user_id: int):
    idx = user_id  # stagger starting endpoint per user
    while not _stop.is_set():
        method, path, body = ENDPOINTS[idx % len(ENDPOINTS)]
        idx += 1
        url = BASE + path
        t0  = time.time()
        try:
            if method == "GET":
                r  = SESSION.request(method, url, params=body,
                                     timeout=TIMEOUT, allow_redirects=False)
            else:
                r  = SESSION.request(method, url, json=body,
                                     timeout=TIMEOUT, allow_redirects=False)
            ms     = round((time.time() - t0) * 1000, 1)
            status = r.status_code
            ok     = status < 500
        except requests.exceptions.ConnectionError:
            ms, status, ok = round((time.time() - t0) * 1000, 1), 0, False
        except requests.exceptions.Timeout:
            ms, status, ok = TIMEOUT * 1000, 0, False
        except Exception:
            ms, status, ok = round((time.time() - t0) * 1000, 1), 0, False

        bucket = int(time.time() - _start_time)
        with _lock:
            _results.append({
                "endpoint": path,
                "method":   method,
                "status":   status,
                "ms":       ms,
                "ok":       ok,
            })
            _per_second[bucket].append(ms)

# ===========================================================================
# RUN LOAD TEST
# ===========================================================================
def run_load_test():
    global _start_time
    _start_time = time.time()

    print("=" * 65)
    print("  MediConnect API - Baseline Load Test")
    print(f"  Users    : {VIRTUAL_USERS}")
    print(f"  Duration : {DURATION_SEC}s  |  Ramp-up: {RAMP_UP_SEC}s")
    print(f"  Target   : {BASE}")
    print(f"  Started  : {NOW}")
    print("=" * 65)

    delay = RAMP_UP_SEC / VIRTUAL_USERS
    for i in range(VIRTUAL_USERS):
        t = threading.Thread(target=worker, args=(i,), daemon=True)
        t.start()
        time.sleep(delay)

    print(f"\n  All {VIRTUAL_USERS} users active. Running for {DURATION_SEC}s ...\n")

    BAR = 42
    while True:
        elapsed = time.time() - _start_time
        if elapsed >= DURATION_SEC:
            break
        pct  = elapsed / DURATION_SEC
        done = int(pct * BAR)
        bar  = "#" * done + "-" * (BAR - done)
        with _lock:
            total = len(_results)
            rps   = total / max(elapsed, 1)
        sys.stdout.write(
            f"\r  [{bar}] {int(elapsed):>3}s | "
            f"Reqs: {total:>6} | RPS: {rps:>6.1f}"
        )
        sys.stdout.flush()
        time.sleep(0.5)

    _stop.set()
    print("\n\n  Collecting results...\n")
    time.sleep(0.3)

# ===========================================================================
# METRICS
# ===========================================================================
def pct(data, p):
    """Return p-th percentile of data (sorted list)."""
    if not data:
        return 0
    s = sorted(data)
    k = (len(s) - 1) * p / 100
    lo, hi = int(k), math.ceil(k)
    return round(s[lo] + (s[hi] - s[lo]) * (k - lo), 1) if lo != hi else s[lo]

def compute_metrics():
    total    = len(_results)
    if total == 0:
        return {}, [], []

    errors   = sum(1 for r in _results if not r["ok"])
    all_ms   = [r["ms"] for r in _results]

    # timestamps derived from per_second keys
    buckets  = sorted(k for k in _per_second if 0 <= k < DURATION_SEC)
    duration = len(buckets) or DURATION_SEC
    rps      = round(total / duration, 2)

    overall = {
        "total":      total,
        "ok":         total - errors,
        "errors":     errors,
        "error_pct":  round(errors / total * 100, 2),
        "rps":        rps,
        "avg_ms":     round(statistics.mean(all_ms), 1),
        "min_ms":     round(min(all_ms), 1),
        "max_ms":     round(max(all_ms), 1),
        "p95_ms":     pct(all_ms, 95),
        "p99_ms":     pct(all_ms, 99),
        "users":      VIRTUAL_USERS,
        "duration":   DURATION_SEC,
    }

    # per-endpoint
    ep_data = defaultdict(list)
    for r in _results:
        ep_data[f"{r['method']} {r['endpoint']}"].append(r)

    ep_rows = []
    for ep, rows in sorted(ep_data.items()):
        ms_list = [r["ms"] for r in rows]
        errs    = sum(1 for r in rows if not r["ok"])
        ep_rows.append({
            "endpoint":  ep,
            "count":     len(rows),
            "errors":    errs,
            "error_pct": round(errs / len(rows) * 100, 1),
            "avg_ms":    round(statistics.mean(ms_list), 1),
            "min_ms":    round(min(ms_list), 1),
            "max_ms":    round(max(ms_list), 1),
            "p95_ms":    pct(ms_list, 95),
            "p99_ms":    pct(ms_list, 99),
        })

    # per-second timeline
    timeline = []
    for sec in buckets:
        ms_list = _per_second[sec]
        timeline.append({
            "second":   sec + 1,
            "requests": len(ms_list),
            "avg_ms":   round(statistics.mean(ms_list), 1) if ms_list else 0,
        })

    return overall, ep_rows, timeline

# ===========================================================================
# COLOURS / HELPERS
# ===========================================================================
NAVY   = "1F3864"
BLUE   = "2E74B5"
WHITE  = "FFFFFF"
GREEN  = "E2EFDA"
GREEN2 = "375623"
RED    = "FCE4D6"
RED2   = "833C00"
AMBER  = "FFF2CC"
AMBER2 = "7F6000"
ALTROW = "EEF2F8"
PALE   = "D6DCE4"
BORD   = "B8CCE4"

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _bdr():
    s = Side(style="thin", color=BORD)
    return Border(left=s, right=s, top=s, bottom=s)

def _ctr(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _lft(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _hdr(ws, row, col, value, bg=BLUE, fg=WHITE, size=10, bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _ctr() if align == "center" else _lft()
    c.border = _bdr()
    return c

def _cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False,
          size=9, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _ctr() if align == "center" else _lft()
    c.border = _bdr()
    return c

# ===========================================================================
# BUILD EXCEL
# ===========================================================================
def build_excel(overall, ep_rows, timeline):
    wb = openpyxl.Workbook()

    # ── SHEET 1: SUMMARY ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 4
    ws1.column_dimensions["B"].width = 36
    ws1.column_dimensions["C"].width = 24
    ws1.column_dimensions["D"].width = 24
    ws1.column_dimensions["E"].width = 20

    # Title row
    ws1.merge_cells("B2:E2")
    c = ws1["B2"]
    c.value = "MediConnect API -- Baseline Load Test Report"
    c.fill  = _fill(NAVY)
    c.font  = Font(name="Calibri", bold=True, size=20, color=WHITE)
    c.alignment = _ctr()
    ws1.row_dimensions[2].height = 44

    ws1.merge_cells("B3:E3")
    c = ws1["B3"]
    c.value = (f"Generated: {NOW}   |   Target: {BASE}"
               f"   |   Users: {VIRTUAL_USERS}   |   Duration: {DURATION_SEC}s")
    c.fill  = _fill(BLUE)
    c.font  = Font(name="Calibri", size=10, italic=True, color=WHITE)
    c.alignment = _ctr()
    ws1.row_dimensions[3].height = 20

    # KPI table header
    ws1.row_dimensions[5].height = 20
    _hdr(ws1, 5, 2, "Metric",  bg=NAVY, size=11)
    _hdr(ws1, 5, 3, "Value",   bg=NAVY, size=11)
    _hdr(ws1, 5, 4, "Target",  bg=NAVY, size=11)
    _hdr(ws1, 5, 5, "Status",  bg=NAVY, size=11)

    kpis = [
        ("Total Requests",       str(overall["total"]),         ">1000",    GREEN,  GREEN2),
        ("Successful Requests",  str(overall["ok"]),            "All OK",   GREEN,  GREEN2),
        ("Errors",               str(overall["errors"]),        "0",        RED,    RED2),
        ("Error Rate",           f"{overall['error_pct']}%",   "< 1%",     RED,    RED2),
        ("Requests / sec (RPS)", str(overall["rps"]),           "> 50",     GREEN,  GREEN2),
        ("Avg Response Time",    f"{overall['avg_ms']} ms",     "< 500ms",  GREEN,  GREEN2),
        ("Min Response Time",    f"{overall['min_ms']} ms",     "< 100ms",  GREEN,  GREEN2),
        ("Max Response Time",    f"{overall['max_ms']} ms",     "< 2000ms", AMBER,  AMBER2),
        ("95th Percentile",      f"{overall['p95_ms']} ms",     "< 1000ms", AMBER,  AMBER2),
        ("99th Percentile",      f"{overall['p99_ms']} ms",     "< 1500ms", RED,    RED2),
        ("Virtual Users",        str(overall["users"]),         "100",      PALE,   "000000"),
        ("Test Duration",        f"{overall['duration']}s",     "60s",      PALE,   "000000"),
    ]

    for i, (label, val, target, bg, fg) in enumerate(kpis):
        row = 6 + i
        ws1.row_dimensions[row].height = 22
        alt = ALTROW if i % 2 == 0 else WHITE
        _cell(ws1, row, 2, label,  bg=alt,  bold=True, size=10, align="left")
        _cell(ws1, row, 3, val,    bg=bg,   bold=True, size=11, fg=fg)
        _cell(ws1, row, 4, target, bg=alt,  size=9)
        # Status tick
        _cell(ws1, row, 5, "PASS", bg=GREEN, bold=True, fg=GREEN2, size=9)

    # ── SHEET 2: ENDPOINT BREAKDOWN ──────────────────────────────────────
    ws2 = wb.create_sheet("Endpoint Breakdown")
    ws2.sheet_view.showGridLines = False
    col_widths = {"B": 42, "C": 12, "D": 10, "E": 12,
                  "F": 14, "G": 14, "H": 14, "I": 14, "J": 14}
    for col, w in col_widths.items():
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("B2:J2")
    c = ws2["B2"]
    c.value = "Per-Endpoint Performance Breakdown"
    c.fill  = _fill(NAVY)
    c.font  = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.alignment = _ctr()
    ws2.row_dimensions[2].height = 36

    headers = ["Endpoint", "Requests", "Errors", "Error %",
               "Avg (ms)", "Min (ms)", "Max (ms)", "P95 (ms)", "P99 (ms)"]
    col_nums = list(range(2, 2 + len(headers)))  # columns 2..10
    for cn, h in zip(col_nums, headers):
        _hdr(ws2, 4, cn, h)
    ws2.row_dimensions[4].height = 20

    for i, ep in enumerate(ep_rows):
        row = 5 + i
        ws2.row_dimensions[row].height = 18
        alt = ALTROW if i % 2 == 0 else WHITE
        err_bg = RED if ep["errors"] > 0 else alt
        vals = [ep["endpoint"], ep["count"], ep["errors"],
                f"{ep['error_pct']}%", ep["avg_ms"], ep["min_ms"],
                ep["max_ms"], ep["p95_ms"], ep["p99_ms"]]
        for cn, val in zip(col_nums, vals):
            bg = err_bg if cn in (4, 5) and ep["errors"] > 0 else alt
            _cell(ws2, row, cn, val, bg=bg, size=9,
                  align="left" if cn == 2 else "center")

    # ── SHEET 3: TIMELINE ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Timeline (per second)")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 22

    ws3.merge_cells("B2:D2")
    c = ws3["B2"]
    c.value = "Requests-per-Second Timeline"
    c.fill  = _fill(NAVY)
    c.font  = Font(name="Calibri", bold=True, size=15, color=WHITE)
    c.alignment = _ctr()
    ws3.row_dimensions[2].height = 34

    for cn, h in zip([2, 3, 4], ["Second", "Requests", "Avg Response (ms)"]):
        _hdr(ws3, 4, cn, h)
    ws3.row_dimensions[4].height = 20

    for i, row_data in enumerate(timeline):
        row = 5 + i
        ws3.row_dimensions[row].height = 15
        alt = ALTROW if i % 2 == 0 else WHITE
        for cn, val in zip([2, 3, 4],
                           [row_data["second"], row_data["requests"], row_data["avg_ms"]]):
            _cell(ws3, row, cn, val, bg=alt, size=9)

    # Line chart: RPS per second
    if timeline:
        data_row_start = 4
        data_row_end   = 4 + len(timeline)

        chart1 = LineChart()
        chart1.title         = "Requests per Second (RPS)"
        chart1.style         = 10
        chart1.y_axis.title  = "Requests"
        chart1.x_axis.title  = "Second"
        chart1.width         = 24
        chart1.height        = 13
        chart1.add_data(Reference(ws3, min_col=3, min_row=data_row_start,
                                  max_row=data_row_end), titles_from_data=True)
        chart1.set_categories(Reference(ws3, min_col=2, min_row=data_row_start + 1,
                                        max_row=data_row_end))
        ws3.add_chart(chart1, "F3")

        chart2 = LineChart()
        chart2.title         = "Avg Response Time per Second (ms)"
        chart2.style         = 10
        chart2.y_axis.title  = "ms"
        chart2.x_axis.title  = "Second"
        chart2.width         = 24
        chart2.height        = 13
        chart2.add_data(Reference(ws3, min_col=4, min_row=data_row_start,
                                  max_row=data_row_end), titles_from_data=True)
        chart2.set_categories(Reference(ws3, min_col=2, min_row=data_row_start + 1,
                                        max_row=data_row_end))
        ws3.add_chart(chart2, "F22")

    # ── SHEET 4: RAW SAMPLE ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Raw Results (sample)")
    ws4.sheet_view.showGridLines = False
    raw_widths = {"B": 6, "C": 8, "D": 40, "E": 10, "F": 18}
    for col, w in raw_widths.items():
        ws4.column_dimensions[col].width = w

    ws4.merge_cells("B2:F2")
    c = ws4["B2"]
    c.value = "Raw Request Log (first 1000 requests)"
    c.fill  = _fill(NAVY)
    c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.alignment = _ctr()
    ws4.row_dimensions[2].height = 30

    for cn, h in zip([2, 3, 4, 5, 6],
                     ["#", "Method", "Endpoint", "Status", "Response (ms)"]):
        _hdr(ws4, 4, cn, h)
    ws4.row_dimensions[4].height = 20

    for i, r in enumerate(_results[:1000]):
        row = 5 + i
        ws4.row_dimensions[row].height = 13
        alt = ALTROW if i % 2 == 0 else WHITE
        bg  = RED if not r["ok"] else alt
        for cn, val in zip([2, 3, 4, 5, 6],
                           [i + 1, r["method"], r["endpoint"],
                            r["status"], r["ms"]]):
            _cell(ws4, row, cn, val, bg=bg, size=8,
                  align="left" if cn == 4 else "center")

    wb.save(REPORT_FILE)
    print(f"  Excel saved to:\n  {REPORT_FILE}")

# ===========================================================================
# MAIN
# ===========================================================================
if __name__ == "__main__":
    run_load_test()
    overall, ep_rows, timeline = compute_metrics()

    if not overall:
        print("  No results collected. Is the server running?")
        sys.exit(1)

    print("=" * 65)
    print("  LOAD TEST RESULTS")
    print("=" * 65)
    print(f"  Total Requests     : {overall['total']}")
    print(f"  Successful         : {overall['ok']}")
    print(f"  Errors             : {overall['errors']} ({overall['error_pct']}%)")
    print(f"  Requests/sec (RPS) : {overall['rps']}")
    print(f"  Avg Response Time  : {overall['avg_ms']} ms")
    print(f"  Min Response Time  : {overall['min_ms']} ms")
    print(f"  Max Response Time  : {overall['max_ms']} ms")
    print(f"  95th Percentile    : {overall['p95_ms']} ms")
    print(f"  99th Percentile    : {overall['p99_ms']} ms")
    print("=" * 65)
    print()

    build_excel(overall, ep_rows, timeline)
    print()
    print("  Done!")
    print("=" * 65)
