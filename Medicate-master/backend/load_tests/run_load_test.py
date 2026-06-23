"""
=============================================================================
  Medicate API — Load Test Runner
  -------------------------------------------------------------------
  This script:
    1. Installs locust (if missing)
    2. Runs the baseline load test (100 users, 1 min) headlessly
    3. Collects CSV stats from Locust
    4. Generates a professional Excel report with:
         - Summary sheet  (RPS, response times, error rate)
         - Details sheet  (per-endpoint breakdown)
         - Charts         (RPS bar, Response-time bar)
  -------------------------------------------------------------------
  Usage:
      python run_load_test.py [--host http://localhost:5000]
=============================================================================
"""

import subprocess
import sys
import os
import csv
import time
import argparse
from datetime import datetime
from pathlib import Path


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE          = Path(__file__).parent
LOCUSTFILE    = HERE / "locustfile.py"
CSV_PREFIX    = HERE / "load_results"          # Locust writes  load_results_*.csv
REPORT_DIR    = HERE / "reports"
REPORT_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Step 1 – ensure dependencies are available
# ---------------------------------------------------------------------------
def _pip_install(*pkgs):
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "--quiet", *pkgs],
        stdout=subprocess.DEVNULL
    )


def ensure_deps():
    print("[SETUP] Checking dependencies …")
    missing = []
    for pkg in ("locust", "openpyxl"):
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[SETUP] Installing: {', '.join(missing)}")
        _pip_install(*missing)
    print("[SETUP] All dependencies ready.\n")


# ---------------------------------------------------------------------------
# Step 2 – run the Locust baseline test
# ---------------------------------------------------------------------------
def run_locust(host: str, users: int = 100, spawn_rate: int = 10, run_time: str = "1m"):
    print("=" * 65)
    print("  MEDICATE API — BASELINE / LOAD TEST")
    print(f"  Host       : {host}")
    print(f"  Users      : {users}")
    print(f"  Spawn rate : {spawn_rate} users/sec")
    print(f"  Duration   : {run_time}")
    print("=" * 65 + "\n")

    cmd = [
        sys.executable, "-m", "locust",
        "-f",          str(LOCUSTFILE),
        "--headless",
        "-u",          str(users),
        "-r",          str(spawn_rate),
        "-t",          run_time,
        "--host",      host,
        "--csv",       str(CSV_PREFIX),
        "--csv-full-history",
        "--only-summary",
    ]

    print("[RUN] Starting Locust …\n")
    start = time.time()
    result = subprocess.run(cmd, cwd=str(HERE))
    elapsed = time.time() - start

    if result.returncode != 0:
        print("\n[WARN] Locust exited with non-zero code. "
              "Results may still be valid if the test completed.\n")
    else:
        print(f"\n[RUN] Test finished in {elapsed:.1f}s\n")

    return elapsed


# ---------------------------------------------------------------------------
# Step 3 – parse Locust CSV files
# ---------------------------------------------------------------------------
def _safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_locust_csvs():
    """
    Returns:
        summary_rows : list[dict]  – aggregated stats per endpoint
        history_rows : list[dict]  – per-second history (optional)
    """
    stats_file   = Path(str(CSV_PREFIX) + "_stats.csv")
    history_file = Path(str(CSV_PREFIX) + "_stats_history.csv")

    summary_rows = []
    history_rows = []

    if not stats_file.exists():
        print(f"[WARN] Stats file not found: {stats_file}")
        return summary_rows, history_rows

    with stats_file.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Skip the "Aggregated" total row for per-endpoint table;
            # keep it separately for the summary sheet.
            summary_rows.append(row)

    if history_file.exists():
        with history_file.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                history_rows.append(row)

    return summary_rows, history_rows


# ---------------------------------------------------------------------------
# Step 4 – build the Excel report
# ---------------------------------------------------------------------------
def build_excel_report(summary_rows, history_rows, host, users, run_time, elapsed):
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = REPORT_DIR / f"load_test_report_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    #  Colour palette
    # ------------------------------------------------------------------ #
    DARK_BG      = "1E2A3A"
    ACCENT_BLUE  = "2196F3"
    ACCENT_GREEN = "4CAF50"
    ACCENT_RED   = "F44336"
    ACCENT_AMBER = "FF9800"
    LIGHT_ROW    = "EAF4FB"
    ALT_ROW      = "FFFFFF"
    HEADER_FG    = "FFFFFF"

    def _border(style="thin"):
        s = Side(border_style=style, color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr_cell(ws, row, col, value, bg=DARK_BG, fg=HEADER_FG, bold=True, size=11, wrap=False, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = _border()
        return c

    def _data_cell(ws, row, col, value, bg=ALT_ROW, bold=False, align="center", number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", size=10, bold=bold)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = _border()
        if number_format:
            c.number_format = number_format
        return c

    # ================================================================== #
    #  Sheet 1 — SUMMARY
    # ================================================================== #
    ws_sum = wb.active
    ws_sum.title = "📊 Summary"
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 22

    # ── Title block ─────────────────────────────────────────────────────
    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum["A1"]
    title_cell.value     = "🏥 Medicate API — Load Test Report"
    title_cell.font      = Font(bold=True, size=16, color=HEADER_FG, name="Calibri")
    title_cell.fill      = PatternFill("solid", fgColor=DARK_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    ws_sum.merge_cells("A2:B2")
    sub_cell = ws_sum["A2"]
    sub_cell.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font      = Font(italic=True, size=10, color="AAAAAA", name="Calibri")
    sub_cell.fill      = PatternFill("solid", fgColor="2C3E50")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    # ── Config table ────────────────────────────────────────────────────
    config_rows = [
        ("⚙  Test Configuration", ""),
        ("Target Host",    host),
        ("Virtual Users",  users),
        ("Spawn Rate",     "10 users/sec"),
        ("Test Duration",  run_time),
        ("Actual Elapsed", f"{elapsed:.1f}s"),
    ]
    for i, (k, v) in enumerate(config_rows, start=4):
        bg = DARK_BG if i == 4 else (LIGHT_ROW if i % 2 == 0 else ALT_ROW)
        _hdr_cell(ws_sum, i, 1, k, bg=bg, size=10 if i > 4 else 11, align="left")
        _data_cell(ws_sum, i, 2, v, bg=bg if i == 4 else ALT_ROW, align="left")
    ws_sum.row_dimensions[4].height = 22

    # ── Aggregate stats ─────────────────────────────────────────────────
    agg = next((r for r in summary_rows if r.get("Name", "").strip().lower() == "aggregated"), None)
    if not agg:
        agg = {}

    total_reqs   = _safe_float(agg.get("Request Count",  0))
    total_fails  = _safe_float(agg.get("Failure Count",  0))
    avg_rps      = _safe_float(agg.get("Requests/s",     0))
    avg_resp     = _safe_float(agg.get("Average Response Time", 0))
    min_resp     = _safe_float(agg.get("Min Response Time",     0))
    max_resp     = _safe_float(agg.get("Max Response Time",     0))
    p50          = _safe_float(agg.get("50%",  0))
    p90          = _safe_float(agg.get("90%",  0))
    p95          = _safe_float(agg.get("95%",  0))
    p99          = _safe_float(agg.get("99%",  0))
    fail_pct     = (total_fails / total_reqs * 100) if total_reqs else 0.0
    avg_bytes    = _safe_float(agg.get("Average Content Size", 0))

    metric_rows = [
        ("📈 Performance Metrics", "", DARK_BG),
        ("Total Requests",          int(total_reqs),           LIGHT_ROW),
        ("Total Failures",          int(total_fails),          LIGHT_ROW),
        ("Failure Rate",            f"{fail_pct:.2f}%",        ACCENT_RED if fail_pct > 5 else ACCENT_GREEN),
        ("Requests / Second (RPS)", f"{avg_rps:.1f} req/s",    LIGHT_ROW),
        ("Avg Response Time",       f"{avg_resp:.0f} ms",      ALT_ROW),
        ("Min Response Time",       f"{min_resp:.0f} ms",      LIGHT_ROW),
        ("Max Response Time",       f"{max_resp:.0f} ms",      ALT_ROW),
        ("Median (p50)",            f"{p50:.0f} ms",           LIGHT_ROW),
        ("90th Percentile (p90)",   f"{p90:.0f} ms",           ALT_ROW),
        ("95th Percentile (p95)",   f"{p95:.0f} ms",           LIGHT_ROW),
        ("99th Percentile (p99)",   f"{p99:.0f} ms",           ALT_ROW),
        ("Avg Content Size",        f"{avg_bytes:.0f} bytes",  LIGHT_ROW),
    ]

    start_r = 11
    for i, row_data in enumerate(metric_rows, start=start_r):
        label, value, bg = row_data
        is_header = (i == start_r)
        _hdr_cell(ws_sum, i, 1, label,
                  bg=bg, size=11 if is_header else 10, align="left",
                  fg=HEADER_FG if is_header else "1A1A1A")
        cell = ws_sum.cell(row=i, column=2, value=value)
        cell.font      = Font(name="Calibri", size=10,
                              bold=is_header,
                              color=HEADER_FG if is_header else "1A1A1A")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        ws_sum.row_dimensions[i].height = 20

    # ── SLA verdict ─────────────────────────────────────────────────────
    # Thresholds: Avg RT < 5000 ms (Flask dev server, 100 users)
    #             Error Rate < 5%
    SLA_RT_MS   = 5000
    SLA_FAIL_PC = 5
    sla_row  = start_r + len(metric_rows) + 1
    sla_pass = (avg_resp < SLA_RT_MS and fail_pct < SLA_FAIL_PC)
    sla_text = f"✅  SLA PASS  (Avg RT < {SLA_RT_MS} ms  &  Error Rate < {SLA_FAIL_PC}%)" if sla_pass \
               else f"❌  SLA FAIL  (Avg RT ≥ {SLA_RT_MS} ms  OR  Error Rate ≥ {SLA_FAIL_PC}%)"
    ws_sum.merge_cells(f"A{sla_row}:B{sla_row}")
    sla_cell = ws_sum.cell(row=sla_row, column=1, value=sla_text)
    sla_cell.font      = Font(bold=True, size=12, name="Calibri",
                              color="FFFFFF")
    sla_cell.fill      = PatternFill("solid",
                                      fgColor=ACCENT_GREEN if sla_pass else ACCENT_RED)
    sla_cell.alignment = Alignment(horizontal="center", vertical="center")
    sla_cell.border    = _border()
    ws_sum.row_dimensions[sla_row].height = 28

    # ================================================================== #
    #  Sheet 2 — ENDPOINT DETAILS
    # ================================================================== #
    ws_det = wb.create_sheet("📋 Endpoint Details")
    ws_det.sheet_view.showGridLines = False

    det_cols = [
        ("Endpoint / Name",         28),
        ("Method",                   9),
        ("Requests",                11),
        ("Failures",                11),
        ("Fail %",                  10),
        ("RPS",                     10),
        ("Avg RT (ms)",             13),
        ("Min RT (ms)",             13),
        ("Max RT (ms)",             13),
        ("p50 (ms)",                12),
        ("p90 (ms)",                12),
        ("p95 (ms)",                12),
        ("p99 (ms)",                12),
        ("Avg Size (B)",            14),
    ]
    for ci, (hdr, w) in enumerate(det_cols, 1):
        ws_det.column_dimensions[get_column_letter(ci)].width = w
        _hdr_cell(ws_det, 1, ci, hdr, bg=DARK_BG)
    ws_det.row_dimensions[1].height = 28

    endpoint_rows_for_chart = []  # used later for charts

    row_idx = 2
    for row_data in summary_rows:
        name = row_data.get("Name", "").strip()
        if name.lower() == "aggregated":
            continue

        req   = _safe_float(row_data.get("Request Count",  0))
        fail  = _safe_float(row_data.get("Failure Count",  0))
        fp    = (fail / req * 100) if req else 0.0
        rps   = _safe_float(row_data.get("Requests/s",     0))
        avg   = _safe_float(row_data.get("Average Response Time", 0))
        mn    = _safe_float(row_data.get("Min Response Time",     0))
        mx    = _safe_float(row_data.get("Max Response Time",     0))
        _p50  = _safe_float(row_data.get("50%",  0))
        _p90  = _safe_float(row_data.get("90%",  0))
        _p95  = _safe_float(row_data.get("95%",  0))
        _p99  = _safe_float(row_data.get("99%",  0))
        sz    = _safe_float(row_data.get("Average Content Size", 0))
        meth  = row_data.get("Type", "GET")

        bg = LIGHT_ROW if row_idx % 2 == 0 else ALT_ROW
        fail_bg = ACCENT_RED if fp > 10 else (ACCENT_AMBER if fp > 0 else bg)

        values = [name, meth, int(req), int(fail), f"{fp:.1f}%",
                  f"{rps:.2f}", f"{avg:.0f}", f"{mn:.0f}", f"{mx:.0f}",
                  f"{_p50:.0f}", f"{_p90:.0f}", f"{_p95:.0f}", f"{_p99:.0f}",
                  f"{sz:.0f}"]

        for ci, val in enumerate(values, 1):
            cell_bg = fail_bg if ci == 5 else bg
            _data_cell(ws_det, row_idx, ci, val, bg=cell_bg,
                       align="left" if ci == 1 else "center")

        endpoint_rows_for_chart.append({
            "name": name, "avg": avg, "p90": _p90, "rps": rps, "row": row_idx
        })
        row_idx += 1

    # Aggregated totals footer
    if agg:
        _hdr_cell(ws_det, row_idx, 1, "AGGREGATED TOTAL",
                  bg=ACCENT_BLUE, align="left")
        footer_vals = [
            "", int(total_reqs), int(total_fails), f"{fail_pct:.1f}%",
            f"{avg_rps:.2f}", f"{avg_resp:.0f}", f"{min_resp:.0f}",
            f"{max_resp:.0f}", f"{p50:.0f}", f"{p90:.0f}",
            f"{p95:.0f}", f"{p99:.0f}", f"{avg_bytes:.0f}"
        ]
        for ci, val in enumerate(footer_vals, 2):
            _hdr_cell(ws_det, row_idx, ci, val, bg=ACCENT_BLUE)

    # ================================================================== #
    #  Sheet 3 — CHARTS
    # ================================================================== #
    ws_chart = wb.create_sheet("📈 Charts")
    ws_chart.sheet_view.showGridLines = False

    # Write a small data table for the chart engine to reference
    ws_chart["A1"] = "Endpoint"
    ws_chart["B1"] = "Avg RT (ms)"
    ws_chart["C1"] = "p90 RT (ms)"
    ws_chart["D1"] = "RPS"
    for ci in range(1, 5):
        c = ws_chart.cell(row=1, column=ci)
        c.font = Font(bold=True, color=HEADER_FG, name="Calibri")
        c.fill = PatternFill("solid", fgColor=DARK_BG)
        c.alignment = Alignment(horizontal="center")

    for i, ep in enumerate(endpoint_rows_for_chart[:20], start=2):
        ws_chart.cell(row=i, column=1).value = ep["name"]
        ws_chart.cell(row=i, column=2).value = ep["avg"]
        ws_chart.cell(row=i, column=3).value = ep["p90"]
        ws_chart.cell(row=i, column=4).value = ep["rps"]

    n = len(endpoint_rows_for_chart[:20]) + 1

    # Chart 1: Response Time (Avg + p90)
    chart_rt = BarChart()
    chart_rt.type    = "col"
    chart_rt.title   = "Response Time per Endpoint"
    chart_rt.y_axis.title = "Time (ms)"
    chart_rt.x_axis.title = "Endpoint"
    chart_rt.style   = 10
    chart_rt.width   = 26
    chart_rt.height  = 14

    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=n)
    data_avg = Reference(ws_chart, min_col=2, min_row=1, max_row=n)
    data_p90 = Reference(ws_chart, min_col=3, min_row=1, max_row=n)

    chart_rt.add_data(data_avg, titles_from_data=True)
    chart_rt.add_data(data_p90, titles_from_data=True)
    chart_rt.set_categories(cats)
    chart_rt.series[0].graphicalProperties.solidFill = ACCENT_BLUE
    chart_rt.series[1].graphicalProperties.solidFill = ACCENT_AMBER
    ws_chart.add_chart(chart_rt, "F2")

    # Chart 2: Requests per Second
    chart_rps = BarChart()
    chart_rps.type   = "col"
    chart_rps.title  = "Requests per Second (RPS) per Endpoint"
    chart_rps.y_axis.title = "RPS"
    chart_rps.x_axis.title = "Endpoint"
    chart_rps.style  = 10
    chart_rps.width  = 26
    chart_rps.height = 14

    data_rps = Reference(ws_chart, min_col=4, min_row=1, max_row=n)
    chart_rps.add_data(data_rps, titles_from_data=True)
    chart_rps.set_categories(cats)
    chart_rps.series[0].graphicalProperties.solidFill = ACCENT_GREEN
    ws_chart.add_chart(chart_rps, "F22")

    ws_chart.column_dimensions["A"].width = 35

    # ================================================================== #
    #  Sheet 4 — HISTORY (per-second timeline)
    # ================================================================== #
    if history_rows:
        ws_hist = wb.create_sheet("⏱ History")
        ws_hist.sheet_view.showGridLines = False

        hist_cols = ["Timestamp", "User Count", "RPS", "Avg RT (ms)", "Failures/s"]
        for ci, h in enumerate(hist_cols, 1):
            _hdr_cell(ws_hist, 1, ci, h, bg=DARK_BG)
            ws_hist.column_dimensions[get_column_letter(ci)].width = 18

        for ri, row_data in enumerate(history_rows, start=2):
            bg = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
            vals = [
                row_data.get("Timestamp", ""),
                _safe_float(row_data.get("User count", 0)),
                _safe_float(row_data.get("Requests/s", 0)),
                _safe_float(row_data.get("50%", 0)),      # median as proxy
                _safe_float(row_data.get("Failures/s", 0)),
            ]
            for ci, v in enumerate(vals, 1):
                _data_cell(ws_hist, ri, ci, v, bg=bg)

    # ================================================================== #
    #  Save
    # ================================================================== #
    wb.save(str(report_path))
    return report_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Medicate API — Baseline Load Test Runner"
    )
    parser.add_argument(
        "--host", default="http://localhost:5000",
        help="Base URL of the running Medicate API (default: http://localhost:5000)"
    )
    parser.add_argument(
        "--users", type=int, default=100,
        help="Number of concurrent virtual users (default: 100)"
    )
    parser.add_argument(
        "--spawn-rate", type=int, default=10,
        help="Users spawned per second during ramp-up (default: 10)"
    )
    parser.add_argument(
        "--run-time", default="1m",
        help="Test duration e.g. 1m, 90s, 2m (default: 1m)"
    )
    args = parser.parse_args()

    ensure_deps()

    elapsed = run_locust(
        host=args.host,
        users=args.users,
        spawn_rate=args.spawn_rate,
        run_time=args.run_time
    )

    print("[REPORT] Parsing Locust results …")
    summary_rows, history_rows = parse_locust_csvs()

    if not summary_rows:
        print("\n[WARN] No CSV result data found.")
        print("       Make sure the server was running and Locust could connect.\n")
        return

    print("[REPORT] Building Excel report …")
    report_path = build_excel_report(
        summary_rows, history_rows,
        host=args.host,
        users=args.users,
        run_time=args.run_time,
        elapsed=elapsed
    )

    print("\n" + "=" * 65)
    print("  ✅  LOAD TEST COMPLETE")
    print(f"  Excel report: {report_path}")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
